import datetime
import os
import time

import openpyxl
from openpyxl.styles import Font, PatternFill


# -------------------------------------------------------------------------------------------------
def printDataSummary(allVideos):
    reviewedCount = len([vid for vid in allVideos if vid.get('reviewed')])
    print()
    print('----------------------------------------------------------------------------------------')
    print("PROGRAM STATS SUMMARY")
    print('----------------------------------------------------------------------------------------')
    print("Total Videos: {}".format(len(allVideos)))
    print("Totals Reviewed: {}".format(reviewedCount))
    print("Total Unreviewed: {}".format(len(allVideos)-reviewedCount))


# -------------------------------------------------------------------------------------------------
def printNotReviewedSummary(videos):
    print()
    print('----------------------------------------------------------------------------------------')
    print("VIDEOS PENDING REVIEW")
    print('----------------------------------------------------------------------------------------')
    longestLen = 0
    for videoName in videos:
        length = len(videos[videoName])
        if length > longestLen:
            longestLen = length

    for videoName in videos:
        teamName = videos[videoName]
        spacesCount = longestLen - len(teamName) + 2
        print("\t{}:{}{}".format(teamName, " "*spacesCount, videoName))


# -------------------------------------------------------------------------------------------------
def printFailedSummary(videos):
    if len(videos) == 0:
        return

    print()
    print('----------------------------------------------------------------------------------------')
    print("FAILED PARSING")
    print('----------------------------------------------------------------------------------------')
    for entry in videos:
        print("\t{} {}".format(entry, videos[entry]))


# -------------------------------------------------------------------------------------------------
def printTeamPointsForWeekSummary(teamPointsTemplate, DailyTeamPoints):
    print()
    print('----------------------------------------------------------------------------------------')
    print("TEAM POINTS FOR WEEK")
    print('----------------------------------------------------------------------------------------')
    weeklyResults = teamPointsTemplate
    for teamsDay in DailyTeamPoints:
        for team in DailyTeamPoints[teamsDay]:
            weeklyResults[team] += DailyTeamPoints[teamsDay][team]

    for team in weeklyResults:
        print("\t{}: {}".format(team, weeklyResults[team]))


# -------------------------------------------------------------------------------------------------
def setColumnWidths(sheet):
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value + int(value*0.15)
        sheet.column_dimensions[col].auto_size = True


# -------------------------------------------------------------------------------------------------
def writeHeaderRowExcel(sheet, headerNames, row=1):
    column = 1
    for value in headerNames:
        sheet.cell(row=row, column=column, value=value)
        sheet.cell(row, column).font = Font(bold=True)
        sheet.cell(row, column).fill = PatternFill(start_color='7FFFD4', fill_type='solid')
        column += 1


# -------------------------------------------------------------------------------------------------
def writeFailedParsing(sheet, failedPasingVideos):
    writeHeaderRowExcel(sheet, ['Video Name'])
    row = 2
    for videoName in failedPasingVideos:
        sheet.cell(row=row, column=1, value=videoName)

    setColumnWidths(sheet)


# -------------------------------------------------------------------------------------------------
def writePendingReview(sheet, videosPendingReview):
    writeHeaderRowExcel(sheet, ['Video Name', 'Team'])
    row = 2
    for videoName in videosPendingReview:
        teamName = videosPendingReview[videoName]
        sheet.cell(row=row, column=1, value=videoName)
        sheet.cell(row=row, column=2, value=teamName)
        row += 1

    setColumnWidths(sheet)


# -------------------------------------------------------------------------------------------------
def writeSummaryExcel(sheet, allVideos, daysOfTheWeek, dailyTeamPoints):
    # Stats Summary
    writeHeaderRowExcel(sheet, ['Total Players', 'Total Videos', 'Reviewed', 'Unreviewed'])
    row = 2
    reviewedCount = len([vid for vid in allVideos if vid.get('reviewed')])
    sheet.cell(row=row, column=2, value=len(allVideos))
    sheet.cell(row=row, column=3, value=reviewedCount)
    sheet.cell(row=row, column=4, value=len(allVideos) - reviewedCount)

    # Stats of Participation Per Day
    row = row + 2
    writeHeaderRowExcel(sheet, daysOfTheWeek, row)

    dayCounts = {'monday': 0, 'tuesday': 0, 'wednesday': 0, 'thursday': 0, 'friday': 0, 'saturday': 0, 'sunday': 0}
    players = set()
    for video in allVideos:
        videoDay = video.get('day')
        player = video.get('player')
        team = video.get('team')
        players.add('{}-{}'.format(team,player))
        dayCount = dayCounts.get(videoDay.lower())
        dayCounts[videoDay.lower()] = dayCount + 1

    # part of the top summary block, adding in the total number of players
    sheet.cell(row=2, column=1, value=len(players))

    row = row + 1
    column = 1
    for day in dayCounts:
        count = dayCounts[day]
        percentDay = int(count / len(players) * 100)
        sheet.cell(row=row, column=column, value='{} ({}%)'.format(count, percentDay))
        column += 1

    setColumnWidths(sheet)


# -------------------------------------------------------------------------------------------------
def writeSummaryCsv(allVideos, outputPath, fileNameBase):
    import csv
    csvFieldNames = ['name', 'id', 'player', 'team', 'createdTime', 'emails', 'day', 'reviewed']
    csv_file = "{}/{}.csv".format(outputPath, fileNameBase)
    try:
        with open(csv_file, 'w') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=csvFieldNames)
            writer.writeheader()
            for data in allVideos:
                writer.writerow(data)
    except IOError:
        print("I/O error")


# -------------------------------------------------------------------------------------------------
# Unique list of players, emails, etc
def writeChimpFinalExcel(sheet, chimpList):
    playerKeys = set()
    uniqueChimpList = []
    for video in chimpList:
        player = video.get('player')
        team = video.get('team')
        playerKey = '{}-{}'.format(player, team)

        if playerKey in playerKeys:
            continue

        playerKeys.add(playerKey)
        uniqueChimpList.append(video)

    writeDataExcel(sheet, uniqueChimpList)


# -------------------------------------------------------------------------------------------------
# Report on which videos were uploaded on a different day from when they were recorded
def writeUploadDateMismatch(sheet, allVideos, daysOfWeekNames):
    diffUploadDays = []
    for video in allVideos:
        day = video.get('day')
        createdTime = video.get('createdTime')

        dstHours = 7 if time.localtime().tm_isdst else 8
        datetimeObjCreateTime = datetime.datetime.strptime(createdTime, '%Y-%m-%dT%H:%M:%S.%fZ')  # '2020-05-30T03:41:25.012Z'
        pacificCreateTime = datetimeObjCreateTime - datetime.timedelta(hours=dstHours)
        createdDayOfWeek = daysOfWeekNames[pacificCreateTime.weekday()]

        if day and (day.lower() != createdDayOfWeek.lower()):
            videoName = video.get('name')
            player = video.get('player')
            teamName = video.get('team')
            videoDetails = {'name': videoName, 'player': player, 'team': teamName, 'createdTime': pacificCreateTime, 'day': day, 'uploadDay': createdDayOfWeek.lower()}
            diffUploadDays.append(videoDetails)

    headerNames = ['Video Name', 'Player', 'Team', 'Video Created Time', 'Video Day', 'Upload Day']
    writeDataExcel(sheet, diffUploadDays, headerNames)


# -------------------------------------------------------------------------------------------------
def writeDataExcel(sheet, allVideos, headerNames=None):
    defaultHeaderNames = ['Video Name', 'Video Id', 'Player', 'Team', 'Video Created Time', 'Owner Emails', 'Video Day', 'Video Reviewed']
    headerNames = defaultHeaderNames if headerNames is None else headerNames
    writeHeaderRowExcel(sheet, headerNames)

    # openpyxl does things based on 1 instead of 0
    row = 2
    for entry in allVideos:
        column = 1
        for key, values in entry.items():
            # Put the key in the first column for each key in the dictionary
            if key == 'emails':
                values = ','.join(values)
            sheet.cell(row=row, column=column, value=values)
            column += 1
        row += 1

    setColumnWidths(sheet)


# -------------------------------------------------------------------------------------------------
def createReport(allVideos, chimpList, videosPendingReview, videosNeedFixing, weekNumber: int, dayOfWeekNames, dailyTeamPoints):
    now = datetime.datetime.now()
    timestamp = "{}.{}.{}.{}.{}.{}".format(now.year, now.month, now.day, now.hour, now.minute, now.second)
    outputPath = os.path.join('output', timestamp)
    os.makedirs(outputPath)
    fileNameBase = "VanCity Pro Week {} Report".format(weekNumber)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    writeSummaryExcel(sheet, allVideos, dayOfWeekNames, dailyTeamPoints)

    sheetName = 'All Videos'
    workbook.create_sheet(sheetName)
    sheet = workbook[sheetName]
    writeDataExcel(sheet, allVideos)

    sheetName = 'Pending Review'
    workbook.create_sheet(sheetName)
    sheet = workbook[sheetName]
    writePendingReview(sheet, videosPendingReview)

    sheetName = 'Failed Parsing'
    workbook.create_sheet(sheetName)
    sheet = workbook[sheetName]
    writeFailedParsing(sheet, videosNeedFixing)

    sheetName = 'CHIMP Awards'
    workbook.create_sheet(sheetName)
    sheet = workbook[sheetName]
    writeChimpFinalExcel(sheet, chimpList)

    sheetName = 'Wrong Upload Day'
    workbook.create_sheet(sheetName)
    sheet = workbook[sheetName]
    writeUploadDateMismatch(sheet, allVideos, dayOfWeekNames)

    workbook.save(filename="{}/{}.xlsx".format(outputPath, fileNameBase))
